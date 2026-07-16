from io import BytesIO

from openpyxl import load_workbook

from app.excel_export import ExportRecord, build_excel
from app.models import EmployeeInfo, MethodScores, ParsedProfile, PsychologicalReport


def test_excel_has_expected_columns_and_report():
    profile = ParsedProfile(
        employee=EmployeeInfo(full_name="TEST_SUBJECT_001", position="Руководитель", department="TEST_ORG"),
        methods=MethodScores(),
    )
    report = PsychologicalReport(
        emotional_motivation="Мотивация",
        management_style="Управление",
        communication_style="Коммуникация",
        risk_factors="Риски",
        recommendations=[f"Рекомендация {n}" for n in range(1, 11)],
    )
    content = build_excel([ExportRecord(profile=profile, report=report)])
    wb = load_workbook(BytesIO(content))
    ws = wb["Характеристики"]
    assert ws.max_column == 8
    assert ws["B2"].value.startswith("TEST_SUBJECT_001")
    assert ws["C2"].value == "Мотивация"
    assert "10. Рекомендация 10" in ws["G2"].value
    assert wb["Сведения"].sheet_state == "hidden"


def test_excel_escapes_formula_injection():
    profile = ParsedProfile(employee=EmployeeInfo(full_name="=HYPERLINK(\"bad\")"), methods=MethodScores())
    report = PsychologicalReport(
        emotional_motivation="=1+1", management_style="текст", communication_style="текст", risk_factors="текст",
        recommendations=[f"Рекомендация {n}" for n in range(1, 11)],
    )
    wb = load_workbook(BytesIO(build_excel([ExportRecord(profile=profile, report=report)])), data_only=False)
    assert wb["Характеристики"]["B2"].data_type == "s"
    assert wb["Характеристики"]["C2"].data_type == "s"
