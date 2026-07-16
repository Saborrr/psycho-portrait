from pathlib import Path

from fastapi.testclient import TestClient
from io import BytesIO
from openpyxl import load_workbook

from app.main import app
from app.models import PsychologicalReport


def test_api_requires_key(monkeypatch):
    monkeypatch.setenv("APP_API_KEY", "test-secret")
    monkeypatch.setenv("ALLOW_INSECURE_LOCALHOST", "false")
    client = TestClient(app)
    assert client.get("/health").status_code == 200
    assert client.get("/api/settings").status_code == 401
    response = client.get("/api/settings", headers={"X-API-Key": "test-secret"})
    assert response.status_code == 200
    assert response.headers["cache-control"].startswith("no-store")
    assert response.headers["x-content-type-options"] == "nosniff"


def test_pwa_shell_is_available_without_caching_api(monkeypatch):
    client = TestClient(app)
    assert client.get("/").status_code == 200
    worker = client.get("/sw.js")
    assert worker.status_code == 200
    assert worker.headers["service-worker-allowed"] == "/"
    assert 'url.pathname.startsWith("/api/")' in worker.text
    manifest = client.get("/static/manifest.webmanifest").json()
    assert manifest["display"] == "standalone"
    assert {icon["sizes"] for icon in manifest["icons"]} == {"192x192", "512x512"}


def test_parse_endpoint_accepts_real_pptx_container(monkeypatch):
    monkeypatch.setenv("APP_API_KEY", "test-secret")
    monkeypatch.setenv("STORE_HISTORY", "false")
    client = TestClient(app)
    path = Path(__file__).parent.parent / "samples" / "sample_efko_full.pptx"
    with path.open("rb") as handle:
        response = client.post(
            "/api/parse",
            headers={"X-API-Key": "test-secret"},
            files={"file": (path.name, handle, "application/vnd.openxmlformats-officedocument.presentationml.presentation")},
        )
    assert response.status_code == 200, response.text
    body = response.json()
    assert "raw_text" not in body["profile"]
    assert body["profile"]["employee"]["extra"] == {}


def test_batch_endpoint_returns_real_xlsx(monkeypatch):
    monkeypatch.setenv("APP_API_KEY", "test-secret")
    monkeypatch.setenv("STORE_HISTORY", "false")

    async def fake_report(_profile):
        text = " ".join(["Проверяемая рабочая гипотеза описывает поведение сотрудника в деловых ситуациях без категоричных кадровых выводов."] * 12)
        return PsychologicalReport(
            emotional_motivation=text, management_style=text, communication_style=text, risk_factors=text,
            recommendations=[f"Практический совет {n}" for n in range(1, 11)],
        )

    monkeypatch.setattr("app.main.generate_report", fake_report)
    client = TestClient(app)
    path = Path(__file__).parent.parent / "samples" / "sample_efko_full.pptx"
    data = path.read_bytes()
    response = client.post(
        "/api/batch/generate-xlsx",
        headers={"X-API-Key": "test-secret"},
        files=[
            ("files", ("employee-1.pptx", data, "application/vnd.openxmlformats-officedocument.presentationml.presentation")),
            ("files", ("employee-2.pptx", data, "application/vnd.openxmlformats-officedocument.presentationml.presentation")),
        ],
    )
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    assert workbook["Характеристики"].max_row == 3
