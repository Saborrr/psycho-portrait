"""Формирование итоговой книги Excel с характеристиками сотрудников."""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import ParsedProfile, PsychologicalReport
from .prompts import PROMPT_VERSION


@dataclass
class ExportRecord:
    profile: ParsedProfile
    report: PsychologicalReport
    photo: tuple[bytes, str] | None = None


HEADERS = (
    "№",
    "Сотрудник",
    "Эмоциональная природа мотивации",
    "Стиль управления",
    "Стиль коммуникации",
    "Факторы риска",
    "Рекомендации",
    "Фото",
)


def _employee_cell(profile: ParsedProfile) -> str:
    values = [profile.employee.full_name or "Сотрудник"]
    if profile.employee.department:
        values.append(profile.employee.department)
    if profile.employee.position:
        values.append(profile.employee.position)
    return _excel_safe("\n".join(values))


def _excel_safe(value: str) -> str:
    """Не даем тексту из документа или LLM превратиться в формулу Excel."""
    cleaned = value.replace("\x00", "")
    if cleaned.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + cleaned
    return cleaned


def build_excel(records: list[ExportRecord]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Характеристики"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(1, len(records) + 1)}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 38

    widths = (6, 30, 47, 47, 47, 47, 43, 24)
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    image_buffers: list[BytesIO] = []
    for index, record in enumerate(records, 1):
        row = index + 1
        report = record.report
        risks = report.risk_factors
        if report.quality_warnings:
            risks += "\n\nОграничения интерпретации: " + " ".join(report.quality_warnings)
        values = (
            index,
            _employee_cell(record.profile),
            _excel_safe(report.emotional_motivation),
            _excel_safe(report.management_style),
            _excel_safe(report.communication_style),
            _excel_safe(risks),
            _excel_safe("\n".join(f"{n}. {item}" for n, item in enumerate(report.recommendations, 1))),
        )
        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.alignment = Alignment(
                horizontal="center" if col == 1 else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = border
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 8).border = border
        ws.row_dimensions[row].height = 230

        if record.photo:
            try:
                buffer = BytesIO(record.photo[0])
                image_buffers.append(buffer)
                image = XLImage(buffer)
                max_width, max_height = 165, 295
                ratio = min(max_width / image.width, max_height / image.height)
                image.width = int(image.width * ratio)
                image.height = int(image.height * ratio)
                image.anchor = f"H{row}"
                ws.add_image(image)
            except Exception:
                ws.cell(row, 8, "Фото не удалось встроить")

    info = wb.create_sheet("Сведения")
    info.append(("Назначение", "Черновик для проверки квалифицированным психологом"))
    info.append(("Ограничение", "Не использовать как единственное основание кадрового решения"))
    info.append(("Версия промпта", PROMPT_VERSION))
    info.append(("Версия методики", records[0].profile.methodology_version if records else "efko-lecture-2025.01"))
    info.sheet_state = "hidden"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
